"""
Management command to bulk-import a residential complex from the Excel template.

Usage:
    python manage.py import_conjunto --file /path/to/conjunto.xlsx [--send-emails]

Excel template sheets (in order):
    1. Conjunto       — datos del conjunto
    2. Torres         — una fila por torre
    3. Administrador  — un administrador
    4. Propietarios   — lista de propietarios
    5. Portería       — lista de portería
"""
import secrets
import string

from django.core.management.base import BaseCommand, CommandError
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from django.conf import settings

try:
    import openpyxl
except ImportError:
    openpyxl = None


def _random_password(length=12):
    alphabet = string.ascii_letters + string.digits + '!@#$%'
    return ''.join(secrets.choice(alphabet) for _ in range(length))


def _send_welcome_email(email, nombre, conjunto_nombre, cedula, password):
    """Envía email de bienvenida con credenciales al nuevo usuario."""
    try:
        context = {
            'nombre': nombre,
            'conjunto_nombre': conjunto_nombre,
            'cedula': cedula,
            'password': password,
            'login_url': getattr(settings, 'SITE_URL', 'https://kislev.net.co') + '/accounts/login/',
        }
        html = render_to_string('emails/bienvenida_credenciales.html', context)
        text = (
            f"Hola {nombre},\n\n"
            f"Bienvenido a {conjunto_nombre} en KislevSmart.\n\n"
            f"Tus credenciales de acceso:\n"
            f"  Usuario (cédula): {cedula}\n"
            f"  Contraseña temporal: {password}\n\n"
            f"Por seguridad deberás cambiarla en tu primer inicio de sesión.\n\n"
            f"Ingresa en: {context['login_url']}"
        )
        msg = EmailMultiAlternatives(
            subject=f'Bienvenido a KislevSmart — {conjunto_nombre}',
            body=text,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[email],
        )
        msg.attach_alternative(html, 'text/html')
        msg.send(fail_silently=False)
        return True
    except Exception as exc:
        return str(exc)


class Command(BaseCommand):
    help = 'Importa un conjunto residencial completo desde el Excel plantilla'

    def add_arguments(self, parser):
        parser.add_argument('--file', required=True, help='Ruta al archivo Excel (.xlsx)')
        parser.add_argument(
            '--send-emails',
            action='store_true',
            default=False,
            help='Envía email de bienvenida con credenciales a cada usuario creado',
        )

    def handle(self, *args, **options):
        if openpyxl is None:
            raise CommandError('openpyxl no está instalado. Ejecuta: pip install openpyxl')

        filepath = options['file']
        send_emails = options['send_emails']

        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
        except FileNotFoundError:
            raise CommandError(f'Archivo no encontrado: {filepath}')
        except Exception as exc:
            raise CommandError(f'No se pudo leer el archivo: {exc}')

        # ── 1. CONJUNTO ──────────────────────────────────────────────────────
        ws = wb['Conjunto']
        data = {row[0].value: row[1].value for row in ws.iter_rows(min_row=2) if row[0].value}

        required_conjunto = ['nombre', 'nit', 'direccion']
        for field in required_conjunto:
            if not data.get(field):
                raise CommandError(f'Hoja "Conjunto": falta el campo "{field}"')

        from accounts.models import ConjuntoResidencial, Torre, Usuario

        tipo_dist = str(data.get('tipo_distribucion', '') or 'torre_apto').strip()
        valid_tipos = [c[0] for c in ConjuntoResidencial.DISTRIBUCION_CHOICES]
        if tipo_dist not in valid_tipos:
            self.stdout.write(self.style.WARNING(
                f'  tipo_distribucion "{tipo_dist}" no válido — usando "torre_apto". '
                f'Opciones: {", ".join(valid_tipos)}'
            ))
            tipo_dist = 'torre_apto'

        conjunto, created = ConjuntoResidencial.objects.get_or_create(
            nit=str(data['nit']).strip(),
            defaults={
                'nombre': str(data['nombre']).strip(),
                'direccion': str(data.get('direccion', '')).strip(),
                'telefono': str(data.get('telefono', '') or ''),
                'email_contacto': str(data.get('email_contacto', '') or '') or None,
                'link_pago': str(data.get('link_pago', '') or '') or None,
                'tipo_distribucion': tipo_dist,
            },
        )
        if not created and tipo_dist:
            conjunto.tipo_distribucion = tipo_dist
            conjunto.save(update_fields=['tipo_distribucion'])

        action = 'Creado' if created else 'Ya existía'
        self.stdout.write(f'[Conjunto] {action}: {conjunto.nombre} — distribución: {tipo_dist}')

        # ── 2. AGRUPACIONES (Torres/Interiores/Bloques/Manzanas) ─────────────
        # Soporta tanto la hoja antigua "Torres" como la nueva "Agrupaciones"
        sheet_name = 'Agrupaciones' if 'Agrupaciones' in wb.sheetnames else 'Torres'
        ws_torres = wb[sheet_name]
        torres_map = {}  # nombre_torre → Torre instance
        headers = [c.value for c in next(ws_torres.iter_rows(min_row=1, max_row=1))]

        for row in ws_torres.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            row_data = dict(zip(headers, row))
            nombre_torre = str(row_data.get('nombre', '')).strip()
            if not nombre_torre:
                continue
            torre, t_created = Torre.objects.get_or_create(
                conjunto=conjunto,
                nombre=nombre_torre,
                defaults={
                    'numero_pisos': int(row_data.get('numero_pisos') or 1),
                    'aptos_por_piso': int(row_data.get('aptos_por_piso') or 4),
                },
            )
            torres_map[nombre_torre] = torre
            t_action = 'Creada' if t_created else 'Ya existía'
            self.stdout.write(f'  [Torre] {t_action}: {torre.nombre}')

        # ── Helper para crear usuarios ────────────────────────────────────────
        created_users = 0
        skipped_users = 0
        email_errors = []

        def create_user(row_data, user_type):
            nonlocal created_users, skipped_users
            cedula = str(row_data.get('cedula', '') or '').strip()
            nombre = str(row_data.get('nombre', '') or '').strip()
            email = str(row_data.get('email', '') or '').strip()
            if not cedula or not nombre or not email:
                self.stdout.write(self.style.WARNING(f'  [Skip] Fila incompleta: {row_data}'))
                skipped_users += 1
                return

            if Usuario.objects.filter(cedula=cedula, conjunto=conjunto).exists():
                self.stdout.write(f'  [{user_type}] Ya existe cédula {cedula} en este conjunto, omitido.')
                skipped_users += 1
                return

            password = _random_password()
            # Soporta columnas antiguas (torre/apartamento) y nuevas (agrupacion/unidad)
            torre_nombre = str(row_data.get('agrupacion', '') or row_data.get('torre', '') or '').strip()
            torre = torres_map.get(torre_nombre)
            apartamento = str(row_data.get('unidad', '') or row_data.get('apartamento', '') or '').strip()

            usuario = Usuario.objects.create_user(
                cedula=cedula,
                nombre=nombre,
                email=email,
                password=password,
                conjunto=conjunto,
                user_type=user_type,
                phone_number=str(row_data.get('telefono', '') or ''),
                torre=torre,
                apartamento=apartamento,
                must_change_password=True,
            )
            created_users += 1
            self.stdout.write(self.style.SUCCESS(f'  [{user_type}] Creado: {nombre} ({cedula})'))

            if send_emails:
                result = _send_welcome_email(email, nombre, conjunto.nombre, cedula, password)
                if result is True:
                    self.stdout.write(f'    → Email enviado a {email}')
                else:
                    email_errors.append(f'{email}: {result}')
                    self.stdout.write(self.style.WARNING(f'    → Error email {email}: {result}'))

        # ── 3. ADMINISTRADOR ─────────────────────────────────────────────────
        ws_admin = wb['Administrador']
        headers_admin = [c.value for c in next(ws_admin.iter_rows(min_row=1, max_row=1))]
        for row in ws_admin.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            create_user(dict(zip(headers_admin, row)), 'administrador')

        # ── 4. PROPIETARIOS ──────────────────────────────────────────────────
        ws_prop = wb['Propietarios']
        headers_prop = [c.value for c in next(ws_prop.iter_rows(min_row=1, max_row=1))]
        for row in ws_prop.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            create_user(dict(zip(headers_prop, row)), 'propietario')

        # ── 5. PORTERÍA ──────────────────────────────────────────────────────
        ws_port = wb['Portería']
        headers_port = [c.value for c in next(ws_port.iter_rows(min_row=1, max_row=1))]
        for row in ws_port.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            create_user(dict(zip(headers_port, row)), 'porteria')

        # ── Resumen ──────────────────────────────────────────────────────────
        self.stdout.write('')
        self.stdout.write(self.style.SUCCESS(
            f'Importación completada: {created_users} usuarios creados, {skipped_users} omitidos.'
        ))
        if email_errors:
            self.stdout.write(self.style.ERROR(f'Errores de email ({len(email_errors)}):'))
            for err in email_errors:
                self.stdout.write(f'  {err}')
