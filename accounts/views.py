# accounts/views.py
import os
import secrets
from datetime import timedelta
from django.conf import settings
from django.utils import timezone

_DEFAULT_PASSWORD = os.getenv('DEFAULT_USER_PASSWORD', 'kislev123')
_TOKEN_COOKIE = 'kislev_token'
_TOKEN_DAYS   = 30


def _emit_token_cookie(user, response):
    """Crea un PersistentLoginToken en DB y lo pone como cookie en la respuesta."""
    from .models import PersistentLoginToken
    token_value = secrets.token_urlsafe(32)
    PersistentLoginToken.objects.create(
        user=user,
        token=token_value,
        expires_at=timezone.now() + timedelta(days=_TOKEN_DAYS),
    )
    response.set_cookie(
        _TOKEN_COOKIE,
        token_value,
        max_age=_TOKEN_DAYS * 24 * 60 * 60,
        httponly=True,
        samesite='Lax',
        secure=not settings.DEBUG,
        path='/',
    )
    return response


def _sync_password_hash(cedula: str, hashed: str) -> int:
    """Propaga el mismo hash de contraseña a todos los registros con la misma cédula.
    Devuelve la cantidad de filas actualizadas."""
    from .models import Usuario as _Usuario
    return _Usuario.objects.filter(cedula=cedula).update(password=hashed)


from django.views.decorators.csrf import csrf_protect
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from django.views import View
from django import forms
from django.contrib.auth.views import (
    LogoutView as DjangoLogoutView,
    PasswordResetView,
    PasswordResetConfirmView
)
from django.urls import reverse_lazy
from django.contrib.auth.forms import SetPasswordForm
from django.contrib.auth.decorators import login_required
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.mixins import LoginRequiredMixin
from django.utils.decorators import method_decorator
from django.views.generic import TemplateView
from django.utils.translation import gettext as _
from django.core import signing
import time
from django_ratelimit.decorators import ratelimit
from .utils import role_required
from .forms import LoginForm, SelectConjuntoForm
from .models import Usuario, ConjuntoResidencial
from kislevsmart.models import Novedad, NovedadVista, Visitante, VisitanteVehicular, ConfigParqueadero, ParqueaderoCarro, ParqueaderoMoto
from kislevsmart.utils import calcular_cobro_parqueadero, send_email_async, log_envio as _log_envio_global
from django.utils import timezone as tz


# Mantenemos las vistas existentes
@method_decorator(role_required(['administrador']), name='dispatch')
class VisorAdminView(TemplateView):
    template_name = 'accounts/visor_admin.html'

    def get(self, request, *args, **kwargs):
        user = request.user
        conjunto = user.conjunto
        hoy = tz.localdate()
        visitantes_hoy = Visitante.objects.filter(conjunto=conjunto, fecha_generacion__date=hoy).count()
        residentes = Usuario.objects.filter(conjunto=conjunto, user_type='propietario', is_active=True).count()
        from kislevsmart.models import Reserva
        reservas_pendientes = Reserva.objects.filter(sala__conjunto=conjunto, estado='pendiente').count()
        vehiculos_dentro = VisitanteVehicular.objects.filter(
            conjunto=conjunto,
            ultima_lectura__isnull=False,
            segunda_lectura__isnull=True,
        ).count()
        context = {
            'user': user,
            'visitantes_hoy': visitantes_hoy,
            'residentes': residentes,
            'reservas_pendientes': reservas_pendientes,
            'vehiculos_dentro': vehiculos_dentro,
        }
        return self.render_to_response(context)


@method_decorator([login_required, role_required(['porteria', 'administrador'])], name='dispatch')
class ControlPorteriaView(TemplateView):
    template_name = 'accounts/control_porteria.html'

    def get(self, request, *args, **kwargs):
        user = request.user
        novedades_no_vistas = Novedad.objects.filter(
            conjunto=user.conjunto,
            activa=True,
        ).exclude(vistas__usuario=user).order_by('-created_at')
        count = novedades_no_vistas.count()
        context = {
            'user': user,
            'novedades_no_vistas': novedades_no_vistas,
            'novedades_count': count,
        }
        return self.render_to_response(context)


@method_decorator([login_required, role_required(['propietario', 'administrador'])], name='dispatch')
class ControlpropietarioView(TemplateView):
    template_name = 'accounts/visor_propietario.html'

    def get(self, request, *args, **kwargs):
        user = request.user
        novedades_no_vistas = Novedad.objects.filter(
            conjunto=user.conjunto,
            activa=True,
        ).exclude(vistas__usuario=user).order_by('-created_at')

        # Vehículos visitantes activos del propietario
        vehiculos_qs = VisitanteVehicular.objects.filter(
            conjunto=request.user.conjunto,
            tipo_vehiculo__in=['carro', 'moto'],
            ultima_lectura__isnull=False,
            segunda_lectura__isnull=True,
            email_creador=request.user.email
        ).order_by('-ultima_lectura')

        vehiculos_activos = []
        for v in vehiculos_qs:
            config = ConfigParqueadero.objects.filter(
                conjunto=request.user.conjunto,
                tipo_vehiculo=v.tipo_vehiculo
            ).first()
            valor, mins, en_gracia = calcular_cobro_parqueadero(v.ultima_lectura, config)
            horas = mins // 60
            minutos = mins % 60
            vehiculos_activos.append({
                'placa': v.placa,
                'tipo': v.get_tipo_vehiculo_display(),
                'tiempo_str': f'{horas}h {minutos}m' if horas > 0 else f'{minutos}m',
                'valor': valor,
                'en_gracia': en_gracia,
                'entrada': tz.localtime(v.ultima_lectura).strftime('%H:%M'),
            })

        conjunto_id = user.conjunto_id
        disp_carros = ParqueaderoCarro.get_disponibilidad(conjunto_id)
        disp_motos  = ParqueaderoMoto.get_disponibilidad(conjunto_id)

        context = {
            'user': user,
            'novedades_no_vistas': novedades_no_vistas,
            'novedades_count': novedades_no_vistas.count(),
            'vehiculos_activos': vehiculos_activos,
            'disp_carros': disp_carros,
            'disp_motos': disp_motos,
        }
        return self.render_to_response(context)


@method_decorator(csrf_protect, name='dispatch')
@method_decorator(ratelimit(key='ip', rate='5/m', method='POST', block=True), name='post')
class LoginView(View):
    template_name = 'accounts/login.html'

    def get(self, request):
        if request.user.is_authenticated:
            return self._redirect_by_user_type(request.user)
            
        form = LoginForm()
        return render(request, self.template_name, {'form': form})

    def post(self, request):
        form = LoginForm(request.POST)
        if form.is_valid():
            cedula = form.cleaned_data.get('cedula')
            password = form.cleaned_data.get('password')
            # La sesión siempre expira por inactividad (settings.SESSION_COOKIE_AGE).
            # El auto-login persistente se maneja vía cookie kislev_token.

            # Verificar si la cédula existe en algún conjunto
            conjuntos_usuario = ConjuntoResidencial.objects.filter(
                usuario__cedula=cedula,
                usuario__is_active=True,
                estado=True
            ).distinct()

            if not conjuntos_usuario.exists():
                messages.error(
                    request, 
                    'No existe un usuario con esta cédula en ningún conjunto.'
                )
                return render(request, self.template_name, {'form': form})

            # Verificamos la contraseña contra todos los registros de la cédula —
            # por si los hashes difieren (edge case antes de la primera sincronización).
            user_example = None
            for _u in Usuario.objects.filter(cedula=cedula, is_active=True).order_by('pk'):
                if _u.check_password(password):
                    user_example = _u
                    break

            if not user_example:
                messages.error(
                    request, 
                    'Contraseña incorrecta. Por favor, verifíquela.'
                )
                return render(request, self.template_name, {'form': form})

            # Si hay solo un conjunto, autenticamos directamente
            if conjuntos_usuario.count() == 1:
                conjunto = conjuntos_usuario.first()
                user = authenticate(
                    request,
                    cedula=cedula,
                    conjunto=conjunto,
                    password=password
                )
                
                if user:
                    # Especificamos el backend
                    user.backend = 'accounts.backends.CedulaConjuntoBackend'
                    login(request, user)

                    if request.is_secure():
                        request.session.cookie_secure = True

                    user.save()
                    return _emit_token_cookie(user, self._redirect_by_user_type(user))
                else:
                    messages.error(request, 'Error al autenticar. Por favor, verifique sus credenciales.')
                    return render(request, self.template_name, {'form': form})
            else:
                # Si hay múltiples conjuntos, guardamos la cédula y un token firmado
                # (el token prueba que el usuario ya se autenticó correctamente)
                request.session['login_cedula'] = cedula
                request.session['login_token'] = signing.dumps(
                    {'cedula': cedula, 'ts': time.time()}, salt='kislev-login'
                )
                return redirect('accounts:select_conjunto')
        
        return render(request, self.template_name, {'form': form})

    def _redirect_by_user_type(self, user):
        """Helper method para manejar las redirecciones según el tipo de usuario"""
        if user.is_saas_owner:
            return redirect('accounts:saas_dashboard')
        redirects = {
            'propietario': 'accounts:visor_propietario',
            'administrador': 'accounts:visor_admin',
            'porteria': 'accounts:control_porteria'
        }
        return redirect(redirects.get(user.user_type, 'accounts:login'))


@method_decorator(csrf_protect, name='dispatch')
class SelectConjuntoView(View):
    template_name = 'accounts/select_conjunto.html'

    def _verify_login_token(self, request):
        """Verifica el token de sesión firmado. Retorna cedula o None si inválido."""
        cedula = request.session.get('login_cedula')
        token = request.session.get('login_token')
        if not cedula or not token:
            return None
        try:
            data = signing.loads(token, salt='kislev-login', max_age=300)
            if data.get('cedula') != cedula:
                return None
        except Exception:
            return None
        return cedula

    def get(self, request):
        cedula = self._verify_login_token(request)
        if not cedula:
            messages.error(request, 'Sesión expirada. Por favor, inicie sesión nuevamente.')
            return redirect('accounts:login')

        conjuntos_usuario = ConjuntoResidencial.objects.filter(
            usuario__cedula=cedula,
            usuario__is_active=True,
            estado=True
        ).distinct()

        if not conjuntos_usuario.exists():
            messages.error(request, 'No hay conjuntos disponibles para esta cédula.')
            return redirect('accounts:login')

        if conjuntos_usuario.count() == 1:
            conjunto = conjuntos_usuario.first()
            user = Usuario.objects.filter(cedula=cedula, conjunto=conjunto, is_active=True).first()
            if user:
                user.backend = 'accounts.backends.CedulaConjuntoBackend'
                login(request, user)
                request.session.pop('login_cedula', None)
                request.session.pop('login_token', None)
                return _emit_token_cookie(user, self._redirect_by_user_type(user))
            else:
                messages.error(request, 'Error al autenticar. Credenciales inválidas.')
                return redirect('accounts:login')

        form = SelectConjuntoForm(conjuntos=conjuntos_usuario)
        return render(request, self.template_name, {'form': form})
    
    def post(self, request):
        cedula = self._verify_login_token(request)
        if not cedula:
            messages.error(request, 'Sesión expirada. Por favor, inicie sesión nuevamente.')
            return redirect('accounts:login')

        conjuntos_usuario = ConjuntoResidencial.objects.filter(
            usuario__cedula=cedula,
            usuario__is_active=True,
            estado=True
        ).distinct()

        form = SelectConjuntoForm(conjuntos=conjuntos_usuario, data=request.POST)
        if form.is_valid():
            conjunto = form.cleaned_data['conjunto']
            user = Usuario.objects.filter(cedula=cedula, conjunto=conjunto, is_active=True).first()
            if user:
                user.backend = 'accounts.backends.CedulaConjuntoBackend'
                login(request, user)
                request.session.pop('login_cedula', None)
                request.session.pop('login_token', None)
                return _emit_token_cookie(user, self._redirect_by_user_type(user))
            else:
                messages.error(request, 'Error al autenticar. Credenciales inválidas.')
                return redirect('accounts:login')

        return render(request, self.template_name, {'form': form})

    def _redirect_by_user_type(self, user):
        """Helper method para manejar las redirecciones según el tipo de usuario"""
        if user.is_saas_owner:
            return redirect('accounts:saas_dashboard')
        redirects = {
            'propietario': 'accounts:visor_propietario',
            'administrador': 'accounts:visor_admin',
            'porteria': 'accounts:control_porteria'
        }
        return redirect(redirects.get(user.user_type, 'accounts:login'))


@method_decorator(csrf_protect, name='dispatch')
class LogoutView(DjangoLogoutView):
    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated:
            # Borrar todos los tokens persistentes del usuario
            from .models import PersistentLoginToken
            PersistentLoginToken.objects.filter(user=request.user).delete()
            logout(request)
            messages.success(request, 'Has cerrado sesión correctamente.')

        response = redirect('accounts:login')
        response.delete_cookie(_TOKEN_COOKIE, path='/', samesite='Lax')
        return response

    def get_next_page(self):
        return 'login'


class CustomPasswordResetView(PasswordResetView):
    template_name = 'accounts/reset_password.html'
    email_template_name = 'accounts/password_reset_email.html'
    subject_template_name = 'accounts/password_reset_subject.txt'
    success_url = reverse_lazy('accounts:password_reset_done')

    def form_valid(self, form):
        import threading
        email = form.cleaned_data['email']
        if not Usuario.objects.filter(email=email).exists():
            messages.error(self.request, 'No hay una cuenta asociada a ese correo electrónico.')
            return self.form_invalid(form)

        # Enviar email en hilo separado para no bloquear el worker
        def send_reset_email():
            import logging
            logger = logging.getLogger(__name__)
            try:
                form.save(
                    request=self.request,
                    use_https=self.request.is_secure(),
                    from_email=None,
                    email_template_name=self.email_template_name,
                    subject_template_name=self.subject_template_name,
                    html_email_template_name=self.html_email_template_name,
                    extra_email_context=self.extra_email_context,
                )
                logger.info("Email de reset enviado correctamente")
            except Exception as e:
                logger.error(f"Error enviando email de reset: {e}")

        threading.Thread(target=send_reset_email, daemon=True).start()
        messages.success(self.request, 'Se ha enviado un enlace para restablecer la contraseña al correo electrónico mencionado.')
        from django.shortcuts import redirect
        return redirect(self.success_url)


class CustomSetPasswordForm(SetPasswordForm):
    def clean_new_password1(self):
        new_password1 = self.cleaned_data.get('new_password1')
        
        if len(new_password1) < 8:
            raise forms.ValidationError(_("La contraseña debe tener al menos 8 caracteres."))

        return new_password1

    def clean(self):
        cleaned_data = super().clean()
        new_password1 = cleaned_data.get('new_password1')
        new_password2 = cleaned_data.get('new_password2')

        if new_password1 and new_password2 and new_password1 != new_password2:
            raise forms.ValidationError(_("Las contraseñas no coinciden."))

        return cleaned_data


class CustomPasswordChangeView(LoginRequiredMixin, View):
    login_url = '/accounts/login/'

    def get(self, request, *args, **kwargs):
        form = CustomSetPasswordForm(user=request.user)
        return render(request, 'accounts/cambiar_password.html', {'form': form})

    def post(self, request, *args, **kwargs):
        form = CustomSetPasswordForm(user=request.user, data=request.POST)
        if form.is_valid():
            form.save()
            update_session_auth_hash(request, request.user)
            # Propagar hash a todos los conjuntos de esta cédula
            _sync_password_hash(request.user.cedula, request.user.password)
            messages.success(request, 'Tu contraseña ha sido cambiada con éxito.')
            return redirect('dashboard')
        messages.error(request, 'Por favor, corrige los errores.')
        return render(request, 'accounts/cambiar_password.html', {'form': form})


def _mask_email(email):
    """Enmascara el email para mostrarlo sin exponerlo: j***@gm***.com"""
    local, _, domain = email.partition('@')
    m_local = local[0] + '***' if len(local) > 1 else '***'
    parts = domain.rsplit('.', 1)
    if len(parts) == 2:
        m_domain = (parts[0][:2] + '***') + '.' + parts[1]
    else:
        m_domain = domain[:2] + '***'
    return f'{m_local}@{m_domain}'


def _send_reset_email(usuario, reset_url):
    """Envía el email con el link de restablecimiento de contraseña."""
    from django.core.mail import EmailMultiAlternatives
    subject = 'Restablecer contraseña — Kislev'
    text = (
        f'Hola {usuario.nombre},\n\n'
        f'Recibimos una solicitud para restablecer la contraseña de tu cuenta en Kislev.\n\n'
        f'Haz clic en el siguiente enlace (válido por 30 minutos):\n{reset_url}\n\n'
        f'Si no solicitaste este cambio, ignora este mensaje. Tu contraseña no será modificada.\n\n'
        f'— Equipo Kislev'
    )
    html = f"""<!DOCTYPE html><html><head><meta charset="UTF-8"></head><body style="margin:0;padding:0;background:#f0f2f5;font-family:Arial,sans-serif;">
<table width="100%" cellpadding="0" cellspacing="0" style="background:#f0f2f5;padding:32px 16px">
<tr><td align="center"><table width="100%" style="max-width:520px" cellpadding="0" cellspacing="0">
<tr><td style="background:linear-gradient(135deg,#3b1a6e,#6d28d9,#be185d);border-radius:16px 16px 0 0;padding:32px;text-align:center">
  <p style="margin:0;font-size:24px;font-weight:800;color:#fff">🏠 Kislev</p>
  <p style="margin:6px 0 0;font-size:13px;color:rgba(255,255,255,.85)">Restablecer contraseña</p>
</td></tr>
<tr><td style="background:#fff;padding:32px;border-left:1px solid #e5e7eb;border-right:1px solid #e5e7eb">
  <p style="margin:0 0 8px;font-size:18px;font-weight:700;color:#4c1d95">Hola, {usuario.nombre}</p>
  <p style="margin:0 0 24px;font-size:14px;color:#4b5563;line-height:1.7">
    Recibimos una solicitud para restablecer la contraseña de tu cuenta.<br>
    Haz clic en el botón para crear una nueva (el enlace expira en <strong>30 minutos</strong>).
  </p>
  <table cellpadding="0" cellspacing="0" width="100%"><tr><td align="center">
    <a href="{reset_url}" style="display:inline-block;background:linear-gradient(135deg,#4c1d95,#be185d);color:#ffffff!important;-webkit-text-fill-color:#ffffff;text-decoration:none;padding:14px 40px;border-radius:30px;font-size:15px;font-weight:700">
      Cambiar contraseña &rarr;
    </a>
  </td></tr></table>
  <p style="margin:24px 0 0;font-size:12px;color:#9ca3af;line-height:1.6">
    Si no solicitaste este cambio, ignora este mensaje. Tu contraseña no será modificada.<br>
    Si el botón no funciona, copia este enlace: <span style="color:#7c3ded">{reset_url}</span>
  </p>
</td></tr>
<tr><td style="background:#f9fafb;padding:16px 32px;text-align:center;border:1px solid #e5e7eb;border-top:none;border-radius:0 0 16px 16px">
  <p style="margin:0;font-size:12px;color:#9ca3af">kislev.net.co — {usuario.conjunto.nombre}</p>
</td></tr>
</table></td></tr></table></body></html>"""
    msg = EmailMultiAlternatives(
        subject=subject,
        body=text,
        from_email=settings.DEFAULT_FROM_EMAIL,
        to=[usuario.email],
    )
    msg.attach_alternative(html, 'text/html')
    send_email_async(msg, detalle=f'Reset contraseña → {usuario.email}')
    _log_envio_global('email', conjunto=usuario.conjunto, detalle='Reset contraseña')


class RecuperarPasswordView(View):
    """Paso 1 — usuario ingresa cédula y recibe link por email."""

    @method_decorator(ratelimit(key='ip', rate='6/h', method='POST', block=False))
    def post(self, request):
        if getattr(request, 'limited', False):
            messages.error(request, 'Demasiados intentos. Espera un momento e inténtalo de nuevo.')
            return render(request, 'accounts/recuperar_password.html', {'paso': 1})

        cedula = request.POST.get('cedula', '').strip()
        usuario = Usuario.objects.filter(cedula=cedula, is_active=True).first()

        if usuario and usuario.email:
            try:
                token = signing.dumps(
                    {'uid': usuario.pk, 'ph': usuario.password[-14:]},
                    salt='kislev-pw-reset',
                )
                from django.urls import reverse
                reset_url = request.build_absolute_uri(
                    reverse('accounts:recuperar_confirmar', args=[token])
                )
                _send_reset_email(usuario, reset_url)
            except Exception:
                pass  # No revelar si hubo error — anti-enumeración

        # Siempre mostrar el mismo mensaje (anti-enumeración)
        email_parcial = _mask_email(usuario.email) if usuario and usuario.email else ''
        return render(request, 'accounts/recuperar_password.html', {
            'paso': 'enviado',
            'email_parcial': email_parcial,
        })

    def get(self, request):
        return render(request, 'accounts/recuperar_password.html', {'paso': 1})


class RecuperarPasswordConfirmView(View):
    """Paso 2 — valida el token y permite cambiar la contraseña."""

    def _get_usuario(self, token):
        try:
            data = signing.loads(token, salt='kislev-pw-reset', max_age=1800)
            usuario = Usuario.objects.get(pk=data['uid'], is_active=True)
            if usuario.password[-14:] != data['ph']:
                return None  # Token ya usado (contraseña cambió)
            return usuario
        except (signing.BadSignature, signing.SignatureExpired,
                Usuario.DoesNotExist, KeyError, TypeError):
            return None

    def get(self, request, token):
        usuario = self._get_usuario(token)
        if not usuario:
            return render(request, 'accounts/recuperar_confirmar.html', {'invalido': True})
        return render(request, 'accounts/recuperar_confirmar.html', {
            'token': token, 'nombre': usuario.nombre
        })

    def post(self, request, token):
        usuario = self._get_usuario(token)
        if not usuario:
            return render(request, 'accounts/recuperar_confirmar.html', {'invalido': True})

        p1 = request.POST.get('password1', '')
        p2 = request.POST.get('password2', '')
        if len(p1) < 8:
            messages.error(request, 'La contraseña debe tener al menos 8 caracteres.')
            return render(request, 'accounts/recuperar_confirmar.html', {
                'token': token, 'nombre': usuario.nombre
            })
        if p1 != p2:
            messages.error(request, 'Las contraseñas no coinciden.')
            return render(request, 'accounts/recuperar_confirmar.html', {
                'token': token, 'nombre': usuario.nombre
            })

        usuario.set_password(p1)
        usuario.save()
        # Propagar hash a todos los conjuntos de esta cédula
        _sync_password_hash(usuario.cedula, usuario.password)
        messages.success(request, '¡Contraseña actualizada! Ya puedes iniciar sesión.')
        return redirect('accounts:login')


class PasswordResetDoneView(TemplateView):
    template_name = 'accounts/password_reset_done.html'


# ── Force password change ────────────────────────────────────────────────────

class ForcePasswordChangeView(LoginRequiredMixin, View):
    """Intercept login when must_change_password=True."""
    login_url = '/accounts/login/'

    def get(self, request):
        form = CustomSetPasswordForm(user=request.user)
        return render(request, 'accounts/force_password_change.html', {'form': form})

    def post(self, request):
        form = CustomSetPasswordForm(user=request.user, data=request.POST)
        if form.is_valid():
            form.save()
            update_session_auth_hash(request, request.user)
            # Propagar hash y limpiar must_change_password en todos los conjuntos
            _sync_password_hash(request.user.cedula, request.user.password)
            Usuario.objects.filter(cedula=request.user.cedula).update(must_change_password=False)
            messages.success(request, 'Contraseña actualizada. Bienvenido.')
            redirects = {
                'propietario': 'accounts:visor_propietario',
                'administrador': 'accounts:visor_admin',
                'porteria': 'accounts:control_porteria',
            }
            return redirect(redirects.get(request.user.user_type, 'accounts:login'))
        return render(request, 'accounts/force_password_change.html', {'form': form})


# ── SaaS owner dashboard ─────────────────────────────────────────────────────

def saas_required(view_func):
    """Decorator: only allows is_saas_owner users."""
    from functools import wraps
    from django.http import HttpResponseForbidden

    @wraps(view_func)
    def _wrapped(request, *args, **kwargs):
        if not request.user.is_authenticated or not request.user.is_saas_owner:
            return HttpResponseForbidden('Acceso restringido al propietario del SaaS.')
        return view_func(request, *args, **kwargs)
    return _wrapped


@login_required
@saas_required
def saas_dashboard(request):
    """Super-admin dashboard: list of all residential complexes + global send metrics."""
    from kislevsmart.models import LogEnvio, ConfigGlobal
    from django.utils import timezone
    from django.db.models import Count
    import datetime

    from kislevsmart.utils import uso_almacenamiento_conjunto
    conjuntos = ConjuntoResidencial.objects.all().order_by('nombre')
    stats = []
    total_storage_bytes = 0
    for c in conjuntos:
        uso_bytes = uso_almacenamiento_conjunto(c)
        total_storage_bytes += uso_bytes
        cuota_bytes = (c.cuota_almacenamiento_mb or 2048) * 1024 * 1024
        pct_storage = min(round(uso_bytes * 100 / cuota_bytes), 100) if cuota_bytes else 0
        stats.append({
            'conjunto':     c,
            'propietarios': Usuario.objects.filter(conjunto=c, user_type='propietario', is_active=True).count(),
            'admins':       Usuario.objects.filter(conjunto=c, user_type='administrador', is_active=True).count(),
            'porteria':     Usuario.objects.filter(conjunto=c, user_type='porteria', is_active=True).count(),
            'uso_mb':       round(uso_bytes / 1024 / 1024, 1),
            'cuota_mb':     c.cuota_almacenamiento_mb or 2048,
            'pct_storage':  pct_storage,
        })

    hoy = timezone.localdate()
    cfg = ConfigGlobal.get()

    # ── Selector de mes: recibe ?mes=YYYY-MM, por defecto mes actual ──────────
    mes_param = request.GET.get('mes', '')
    try:
        año, mes_num = int(mes_param[:4]), int(mes_param[5:7])
        mes_inicio = datetime.date(año, mes_num, 1)
    except Exception:
        mes_inicio = hoy.replace(day=1)

    if mes_inicio.month == 12:
        mes_fin = datetime.date(mes_inicio.year + 1, 1, 1) - datetime.timedelta(days=1)
    else:
        mes_fin = datetime.date(mes_inicio.year, mes_inicio.month + 1, 1) - datetime.timedelta(days=1)

    # Últimos 12 meses para el select
    meses_opciones = []
    for i in range(11, -1, -1):
        d = (hoy.replace(day=1) - datetime.timedelta(days=i * 28)).replace(day=1)
        meses_opciones.append({
            'valor': d.strftime('%Y-%m'),
            'label': d.strftime('%B %Y'),
            'selected': d.year == mes_inicio.year and d.month == mes_inicio.month,
        })

    # ── Métricas del mes seleccionado ─────────────────────────────────────────
    emails_mes = LogEnvio.objects.filter(tipo='email',    fecha__date__gte=mes_inicio, fecha__date__lte=mes_fin).count()
    wa_mes     = LogEnvio.objects.filter(tipo='whatsapp', fecha__date__gte=mes_inicio, fecha__date__lte=mes_fin).count()

    top_email = (
        LogEnvio.objects
        .filter(tipo='email', fecha__date__gte=mes_inicio, fecha__date__lte=mes_fin, conjunto__isnull=False)
        .values('conjunto__nombre')
        .annotate(total=Count('id'))
        .order_by('-total')[:5]
    )
    top_wa = (
        LogEnvio.objects
        .filter(tipo='whatsapp', fecha__date__gte=mes_inicio, fecha__date__lte=mes_fin, conjunto__isnull=False)
        .values('conjunto__nombre')
        .annotate(total=Count('id'))
        .order_by('-total')[:5]
    )

    pct_email = min(round(emails_mes * 100 / cfg.limite_emails_mes) if cfg.limite_emails_mes else 0, 100)
    pct_wa    = min(round(wa_mes    * 100 / cfg.limite_whatsapp_mes) if cfg.limite_whatsapp_mes else 0, 100)

    return render(request, 'accounts/saas_dashboard.html', {
        'stats':               stats,
        'emails_mes':          emails_mes,
        'wa_mes':              wa_mes,
        'limite_emails':       cfg.limite_emails_mes,
        'limite_wa':           cfg.limite_whatsapp_mes,
        'pct_email':           pct_email,
        'pct_wa':              pct_wa,
        'top_email':           list(top_email),
        'top_wa':              list(top_wa),
        'mes_label':           mes_inicio.strftime('%B %Y'),
        'meses_opciones':      meses_opciones,
        'total_storage_mb':    round(total_storage_bytes / 1024 / 1024, 1),
        'total_storage_gb':    round(total_storage_bytes / 1024 / 1024 / 1024, 2),
        'r2_free_gb':          10,
        'pct_r2':              min(round(total_storage_bytes * 100 / (10 * 1024**3)), 100),
    })


@login_required
@saas_required
def gestionar_conjunto(request, conjunto_id):
    """Panel hub por conjunto para el SaaS owner."""
    from django.utils import timezone
    from django.db.models import Count, Max
    import datetime

    conjunto = get_object_or_404(ConjuntoResidencial, pk=conjunto_id)

    # ── Siempre muestra el mes actual (sin filtro manual) ────────────────────
    hoy = timezone.localdate()
    mes_inicio = hoy.replace(day=1)

    # ── Métricas de envíos ────────────────────────────────────────────────────
    emails_mes     = 0
    wa_mes         = 0
    pct_email      = 0
    pct_wa         = 0
    historico      = []
    ultimos_envios = []
    limite_emails  = 1000
    limite_wa      = 500
    try:
        from kislevsmart.models import LogEnvio, ConfigGlobal
        emails_mes = LogEnvio.objects.filter(conjunto=conjunto, tipo='email',    fecha__date__gte=mes_inicio).count()
        wa_mes     = LogEnvio.objects.filter(conjunto=conjunto, tipo='whatsapp', fecha__date__gte=mes_inicio).count()

        cfg = ConfigGlobal.get()
        limite_emails = cfg.limite_emails_mes
        limite_wa     = cfg.limite_whatsapp_mes
        pct_email = min(round(emails_mes * 100 / limite_emails) if limite_emails else 0, 100)
        pct_wa    = min(round(wa_mes    * 100 / limite_wa)     if limite_wa     else 0, 100)

        for i in range(5, -1, -1):
            d = (hoy.replace(day=1) - datetime.timedelta(days=i * 28)).replace(day=1)
            if d.month < 12:
                fin = d.replace(month=d.month + 1, day=1) - datetime.timedelta(days=1)
            else:
                fin = d.replace(month=12, day=31)
            historico.append({
                'label':    d.strftime('%b %Y'),
                'emails':   LogEnvio.objects.filter(conjunto=conjunto, tipo='email',    fecha__date__gte=d, fecha__date__lte=fin).count(),
                'whatsapp': LogEnvio.objects.filter(conjunto=conjunto, tipo='whatsapp', fecha__date__gte=d, fecha__date__lte=fin).count(),
            })

        ultimos_envios = list(LogEnvio.objects.filter(conjunto=conjunto).order_by('-fecha')[:20])
    except Exception:
        pass

    # ── Residentes ────────────────────────────────────────────────────────────
    total_res   = Usuario.objects.filter(conjunto=conjunto).count()
    activos_res = Usuario.objects.filter(conjunto=conjunto, is_active=True).count()

    # ── Composición de usuarios ───────────────────────────────────────────────
    usuarios_por_tipo = {'propietario': 0, 'arrendatario': 0, 'administrador': 0, 'porteria': 0}
    try:
        from django.db.models import Count as _Count
        for row in Usuario.objects.filter(conjunto=conjunto, is_active=True).values('user_type').annotate(n=_Count('id')):
            usuarios_por_tipo[row['user_type']] = row['n']
        arrendatarios = Usuario.objects.filter(conjunto=conjunto, is_active=True, user_type='propietario', es_arrendatario=True).count()
        usuarios_por_tipo['arrendatario'] = arrendatarios
        usuarios_por_tipo['propietario']  = max(0, usuarios_por_tipo['propietario'] - arrendatarios)
    except Exception:
        pass

    # ── Actividad general ─────────────────────────────────────────────────────
    visitantes_mes     = 0
    visitantes_veh_mes = 0
    paq_pendientes     = 0
    reservas_mes       = 0
    novedades_mes_count = 0
    pagos_mes_count    = 0
    pagos_mes_total    = 0
    try:
        from kislevsmart.models import Visitante
        visitantes_mes = Visitante.objects.filter(
            conjunto=conjunto,
            fecha_generacion__date__gte=mes_inicio,
        ).count()
    except Exception:
        pass
    try:
        from kislevsmart.models import VisitanteVehicular
        visitantes_veh_mes = VisitanteVehicular.objects.filter(
            conjunto=conjunto,
            fecha_generacion__date__gte=mes_inicio,
        ).count()
    except Exception:
        pass
    try:
        from kislevsmart.models import Paquete
        paq_pendientes = Paquete.objects.filter(conjunto=conjunto, estado='pendiente').count()
    except Exception:
        pass
    try:
        from kislevsmart.models import Reserva
        reservas_mes = Reserva.objects.filter(sala__conjunto=conjunto, created_at__date__gte=mes_inicio).count()
    except Exception:
        pass
    try:
        from kislevsmart.models import Novedad
        novedades_mes_count = Novedad.objects.filter(conjunto=conjunto, created_at__date__gte=mes_inicio, activa=True).count()
    except Exception:
        pass
    try:
        from kislevsmart.models import Pago
        from django.db.models import Sum as _Sum
        pagos_qs = Pago.objects.filter(cuota__conjunto=conjunto, fecha_pago__gte=mes_inicio)
        pagos_mes_count = pagos_qs.count()
        pagos_mes_total = pagos_qs.aggregate(t=_Sum('monto_pagado'))['t'] or 0
    except Exception:
        pass

    ultimo_login = Usuario.objects.filter(conjunto=conjunto, last_login__isnull=False).aggregate(
        ult=Max('last_login')
    )['ult']

    # ── Emails por módulo ─────────────────────────────────────────────────────
    emails_por_modulo = []
    try:
        from kislevsmart.models import LogEnvio as _LogEnvio
        detalles = list(_LogEnvio.objects.filter(
            conjunto=conjunto, tipo='email', fecha__date__gte=mes_inicio
        ).values_list('detalle', flat=True))
        module_counts = {}
        for d in detalles:
            if d.startswith('Bienvenida'):
                mod = 'Bienvenida'
            elif 'QR' in d:
                mod = 'QR Acceso'
            elif d.startswith('Novedad'):
                mod = 'Novedades'
            elif 'masiva' in d.lower():
                mod = 'Comunicados masivos'
            elif 'individual' in d.lower():
                mod = 'Comunicados individuales'
            elif 'Factura' in d:
                mod = 'Facturas'
            else:
                mod = 'Otros'
            module_counts[mod] = module_counts.get(mod, 0) + 1
        emails_por_modulo = sorted(module_counts.items(), key=lambda x: x[1], reverse=True)
    except Exception:
        pass

    # ── Salas más reservadas ──────────────────────────────────────────────────
    salas_top = []
    try:
        from kislevsmart.models import Reserva
        from django.db.models import Count as _Count2
        salas_top = list(
            Reserva.objects.filter(sala__conjunto=conjunto, created_at__date__gte=mes_inicio)
            .values('sala__nombre')
            .annotate(total=_Count2('id'))
            .order_by('-total')[:4]
        )
    except Exception:
        pass

    # ── Salud financiera ──────────────────────────────────────────────────────
    cuotas_activas  = 0
    pct_recaudo     = 0
    deuda_estimada  = 0
    ultima_cuota_nombre = ''
    try:
        from kislevsmart.models import Cuota, Pago
        cuotas_activas = Cuota.objects.filter(conjunto=conjunto, fecha_vencimiento__gte=hoy).count()
        ultima_cuota = Cuota.objects.filter(conjunto=conjunto).order_by('-fecha_vencimiento').first()
        if ultima_cuota:
            ultima_cuota_nombre = ultima_cuota.nombre
            total_prop = Usuario.objects.filter(conjunto=conjunto, user_type='propietario', is_active=True).count()
            if total_prop:
                pagaron = Pago.objects.filter(cuota=ultima_cuota).values('propietario').distinct().count()
                pct_recaudo = min(round(pagaron * 100 / total_prop), 100)
                deuda_estimada = max(0, total_prop - pagaron) * ultima_cuota.monto
    except Exception:
        pass

    import json as _json
    return render(request, 'accounts/gestionar_conjunto.html', {
        'conjunto':            conjunto,
        'mes_label':           hoy.strftime('%B %Y'),
        'emails_mes':          emails_mes,
        'wa_mes':              wa_mes,
        'limite_emails':       limite_emails,
        'limite_wa':           limite_wa,
        'pct_email':           pct_email,
        'pct_wa':              pct_wa,
        'historico':           historico,
        'ultimos_envios':      ultimos_envios,
        'total_res':           total_res,
        'activos_res':         activos_res,
        'visitantes_mes':      visitantes_mes,
        'visitantes_veh_mes':  visitantes_veh_mes,
        'paq_pendientes':      paq_pendientes,
        'reservas_mes':        reservas_mes,
        'novedades_mes_count': novedades_mes_count,
        'pagos_mes_count':     pagos_mes_count,
        'pagos_mes_total':     pagos_mes_total,
        'ultimo_login':        ultimo_login,
        'usuarios_por_tipo':   usuarios_por_tipo,
        'emails_por_modulo':   emails_por_modulo,
        'emails_mod_json':     _json.dumps([{'mod': m, 'n': n} for m, n in emails_por_modulo]),
        'salas_top':           salas_top,
        'cuotas_activas':      cuotas_activas,
        'pct_recaudo':         pct_recaudo,
        'deuda_estimada':      deuda_estimada,
        'ultima_cuota_nombre': ultima_cuota_nombre,
    })


@login_required
@saas_required
def update_config_global(request):
    """AJAX — actualiza los límites globales de SES/Twilio."""
    if request.method != 'POST':
        return JsonResponse({'ok': False}, status=405)
    from kislevsmart.models import ConfigGlobal
    cfg = ConfigGlobal.get()
    try:
        cfg.limite_emails_mes   = int(request.POST.get('limite_emails_mes', cfg.limite_emails_mes))
        cfg.limite_whatsapp_mes = int(request.POST.get('limite_whatsapp_mes', cfg.limite_whatsapp_mes))
        cfg.save()
        return JsonResponse({'ok': True})
    except (ValueError, TypeError) as exc:
        return JsonResponse({'ok': False, 'error': str(exc)}, status=400)


@login_required
@saas_required
def update_conjunto_config(request, conjunto_id):
    """AJAX — actualiza campos de configuración de un conjunto específico."""
    if request.method != 'POST':
        return JsonResponse({'ok': False}, status=405)
    conjunto = get_object_or_404(ConjuntoResidencial, pk=conjunto_id)
    allowed = {'horario_atencion', 'link_pago', 'nombre_agrupacion', 'nombre_unidad', 'telefono', 'email_contacto'}
    updated = []
    for field in allowed:
        if field in request.POST:
            value = request.POST[field].strip()
            setattr(conjunto, field, value if value else '')
            updated.append(field)
    if updated:
        conjunto.save(update_fields=updated)
    return JsonResponse({'ok': True})


@login_required
def update_mi_conjunto(request):
    """AJAX — el administrador del conjunto edita la info de su propio conjunto."""
    if request.method != 'POST':
        return JsonResponse({'ok': False}, status=405)
    if request.user.user_type != 'administrador':
        return JsonResponse({'ok': False, 'error': 'Sin permiso'}, status=403)
    conjunto = request.user.conjunto
    if not conjunto:
        return JsonResponse({'ok': False, 'error': 'Sin conjunto asignado'}, status=400)
    allowed = {'horario_atencion', 'link_pago', 'nombre_agrupacion', 'nombre_unidad', 'telefono', 'email_contacto'}
    updated = []
    for field in allowed:
        if field in request.POST:
            value = request.POST[field].strip()
            setattr(conjunto, field, value if value else '')
            updated.append(field)
    if updated:
        conjunto.save(update_fields=updated)
    return JsonResponse({'ok': True})


# ── Excel template download ──────────────────────────────────────────────────

@login_required
@saas_required
def download_template(request):
    """Generate and return the Excel onboarding template."""
    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        from django.http import HttpResponse
        return HttpResponse('openpyxl no instalado.', status=500)

    from django.http import HttpResponse

    wb = openpyxl.Workbook()

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', fgColor='4F2984')
    center = Alignment(horizontal='center', vertical='center')

    def make_sheet(wb, title, headers, example_row=None, first=False):
        ws = wb.active if first else wb.create_sheet(title=title)
        if first:
            ws.title = title
        ws.row_dimensions[1].height = 22
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            ws.column_dimensions[get_column_letter(col_idx)].width = max(len(header) + 4, 18)
        if example_row:
            for col_idx, val in enumerate(example_row, start=1):
                ws.cell(row=2, column=col_idx, value=val)
        return ws

    # Sheet 1: Conjunto
    make_sheet(wb, 'Conjunto',
               ['campo', 'valor', 'opciones_validas'],
               first=True)
    ws_c = wb['Conjunto']
    fields = [
        ('nombre',             'Conjunto Residencial El Prado', ''),
        ('nit',                '900123456-7',                   ''),
        ('direccion',          'Cra 15 # 80-20, Bogotá',        ''),
        ('telefono',           '3001234567',                    ''),
        ('email_contacto',     'admin@elprado.com',             ''),
        ('link_pago',          'https://pagos.ejemplo.com/elprado', ''),
        ('nombre_agrupacion',  'Torre',
         'Torre | Interior | Bloque | Manzana | (dejar vacío si no hay agrupación)'),
        ('nombre_unidad',      'Apto',
         'Apto | Casa | PH | (cualquier nombre libre)'),
        ('horario_atencion',
         'Lunes a viernes: 8:00 a.m. – 5:00 p.m.\nSábados: 9:00 a.m. – 12:00 m.\nDomingos y festivos: Cerrado',
         'Texto libre. Una línea por día o rango. Se muestra a residentes y en correos.'),
    ]
    # Estilo especial para columna de opciones
    from openpyxl.styles import Font as _Font, PatternFill as _Fill, Alignment as _Align
    note_font = _Font(italic=True, color='888888', size=9)
    for i, row_vals in enumerate(fields, start=2):
        ws_c.cell(row=i, column=1, value=row_vals[0])
        ws_c.cell(row=i, column=2, value=row_vals[1])
        if row_vals[2]:
            cell = ws_c.cell(row=i, column=3, value=row_vals[2])
            cell.font = note_font
    ws_c.column_dimensions['C'].width = 70

    # Sheet 2: Agrupaciones (Torres / Interiores / Bloques / Manzanas)
    make_sheet(wb, 'Agrupaciones',
               ['nombre', 'numero_pisos', 'aptos_por_piso'],
               example_row=['Torre 1', 5, 4])
    # Nota en celda A1
    ws_ag = wb['Agrupaciones']
    ws_ag['A1'].comment = None
    note_row = ws_ag.cell(row=ws_ag.max_row + 2, column=1,
                          value='💡 Pon el nombre exacto de cada agrupación (ej: Torre 1, Interior A, Bloque 3). Deja la hoja vacía si el conjunto no usa agrupaciones.')
    note_row.font = note_font

    # Sheet 3: Administrador
    make_sheet(wb, 'Administrador',
               ['cedula', 'nombre', 'email', 'telefono'],
               example_row=['1020304050', 'Carlos Gómez', 'carlos@elprado.com', '3109876543'])

    # Sheet 4: Propietarios
    make_sheet(wb, 'Propietarios',
               ['cedula', 'nombre', 'email', 'telefono', 'agrupacion', 'unidad'],
               example_row=['52987654', 'María López', 'maria@gmail.com', '3151234567', 'Torre 1', '0101'])
    ws_p = wb['Propietarios']
    ws_p.cell(row=1, column=5).comment = None
    note_p = ws_p.cell(row=ws_p.max_row + 2, column=1,
                       value='💡 "agrupacion" = nombre de la Torre/Interior/Bloque/Manzana. Déjalo vacío si el conjunto no usa agrupaciones.')
    note_p.font = note_font

    # Sheet 5: Portería
    make_sheet(wb, 'Portería',
               ['cedula', 'nombre', 'email', 'telefono'],
               example_row=['80654321', 'Pedro Ramos', 'pedro@elprado.com', '3204567890'])

    # Sheet 6: Parqueadero Carros
    make_sheet(wb, 'Parqueadero Carros',
               ['campo', 'valor', 'descripcion'],
               first=True)
    ws_pc = wb['Parqueadero Carros']
    ws_pc.cell(row=2, column=1, value='total_espacios')
    ws_pc.cell(row=2, column=2, value=20)
    ws_pc.cell(row=2, column=3, value='Número total de espacios para carros en el parqueadero').font = note_font
    ws_pc.column_dimensions['C'].width = 55

    # Sheet 7: Parqueadero Motos
    make_sheet(wb, 'Parqueadero Motos',
               ['campo', 'valor', 'descripcion'],
               first=True)
    ws_pm = wb['Parqueadero Motos']
    ws_pm.cell(row=2, column=1, value='total_espacios')
    ws_pm.cell(row=2, column=2, value=10)
    ws_pm.cell(row=2, column=3, value='Número total de espacios para motos en el parqueadero').font = note_font
    ws_pm.column_dimensions['C'].width = 55

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="plantilla_conjunto.xlsx"'
    return response


# ── Upload / import conjunto via web ────────────────────────────────────────

@login_required
@saas_required
def upload_conjunto(request):
    """Web UI to upload the Excel file and import a new conjunto."""
    if request.method == 'GET':
        return render(request, 'accounts/upload_conjunto.html')

    excel_file = request.FILES.get('excel')
    send_emails = request.POST.get('send_emails') == 'on'
    nombre_agrupacion_form = request.POST.get('nombre_agrupacion', 'Torre').strip()
    nombre_unidad_form = request.POST.get('nombre_unidad', 'Apto').strip() or 'Apto'

    if not excel_file:
        messages.error(request, 'Debes adjuntar un archivo Excel.')
        return render(request, 'accounts/upload_conjunto.html')

    if not excel_file.name.endswith('.xlsx'):
        messages.error(request, 'Solo se aceptan archivos .xlsx')
        return render(request, 'accounts/upload_conjunto.html')

    import io
    import secrets
    import string

    try:
        import openpyxl
    except ImportError:
        messages.error(request, 'openpyxl no está instalado en el servidor.')
        return render(request, 'accounts/upload_conjunto.html')

    from django.core.mail import EmailMultiAlternatives
    from django.template.loader import render_to_string

    def _random_password(length=12):
        alphabet = string.ascii_letters + string.digits + '!@#$%'
        return ''.join(secrets.choice(alphabet) for _ in range(length))

    def _send_welcome(email, nombre, conjunto_nombre, cedula, password):
        try:
            context = {
                'nombre': nombre,
                'conjunto_nombre': conjunto_nombre,
                'cedula': cedula,
                'password': password,
                'login_url': getattr(settings, 'SITE_URL', 'https://kislev.net.co') + '/accounts/login/',
            }
            html = render_to_string('emails/bienvenida_credenciales.html', context)
            text = (
                f"Hola {nombre},\n\nBienvenido a {conjunto_nombre} en Kislev.\n\n"
                f"Usuario: {cedula}\nContraseña temporal: {password}\n\n"
                f"Cámbiala en tu primer inicio de sesión.\n{context['login_url']}"
            )
            msg = EmailMultiAlternatives(
                subject=f'Bienvenido a Kislev — {conjunto_nombre}',
                body=text,
                from_email=settings.DEFAULT_FROM_EMAIL,
                to=[email],
            )
            msg.attach_alternative(html, 'text/html')
            send_email_async(msg, detalle=f'Bienvenida upload → {email}')
            from accounts.models import ConjuntoResidencial as _CR
            _conj = _CR.objects.filter(nombre=conjunto_nombre).first()
            _log_envio_global('email', conjunto=_conj, detalle=f'Bienvenida: {nombre}')
            return True
        except Exception as exc:
            return str(exc)

    try:
        wb = openpyxl.load_workbook(io.BytesIO(excel_file.read()), data_only=True)
    except Exception as exc:
        messages.error(request, f'No se pudo leer el Excel: {exc}')
        return render(request, 'accounts/upload_conjunto.html')

    from django.db import transaction as _tx
    created_count = 0
    skipped_count = 0
    email_errors = []
    conjunto = None

    try:
        with _tx.atomic():
            ws_c = wb['Conjunto']
            data = {row[0].value: row[1].value for row in ws_c.iter_rows(min_row=2) if row[0].value}

            for field in ('nombre', 'nit', 'direccion'):
                if not data.get(field):
                    raise ValueError(f'Hoja "Conjunto": falta el campo "{field}"')

            nombre_agrupacion = str(data.get('nombre_agrupacion', '') or nombre_agrupacion_form).strip()
            nombre_unidad = str(data.get('nombre_unidad', '') or nombre_unidad_form or 'Apto').strip()

            conjunto, _ = ConjuntoResidencial.objects.get_or_create(
                nit=str(data['nit']).strip(),
                defaults={
                    'nombre': str(data['nombre']).strip(),
                    'direccion': str(data.get('direccion', '')).strip(),
                    'telefono': str(data.get('telefono', '') or ''),
                    'email_contacto': str(data.get('email_contacto', '') or '') or None,
                    'link_pago': str(data.get('link_pago', '') or '') or None,
                    'nombre_agrupacion': nombre_agrupacion,
                    'nombre_unidad': nombre_unidad,
                    'horario_atencion': str(data.get('horario_atencion', '') or '').strip(),
                },
            )

            # Agrupaciones — soporta hoja "Agrupaciones" (nueva) y "Torres" (antigua)
            torres_map = {}
            sheet_name = 'Agrupaciones' if 'Agrupaciones' in wb.sheetnames else 'Torres'
            ws_torres = wb[sheet_name]
            headers = [c.value for c in next(ws_torres.iter_rows(min_row=1, max_row=1))]
            for row in ws_torres.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                rd = dict(zip(headers, row))
                nombre_torre = str(rd.get('nombre', '')).strip()
                if not nombre_torre or nombre_torre.startswith('💡'):
                    continue
                from accounts.models import Torre
                torre, _ = Torre.objects.get_or_create(
                    conjunto=conjunto,
                    nombre=nombre_torre,
                    defaults={
                        'numero_pisos': int(rd.get('numero_pisos') or 1),
                        'aptos_por_piso': int(rd.get('aptos_por_piso') or 4),
                    },
                )
                torres_map[nombre_torre] = torre

            def create_user(row_data, user_type):
                nonlocal created_count, skipped_count
                cedula = str(row_data.get('cedula', '') or '').strip()
                nombre = str(row_data.get('nombre', '') or '').strip()
                email = str(row_data.get('email', '') or '').strip()
                if not cedula or not nombre or not email:
                    skipped_count += 1
                    return
                if Usuario.objects.filter(cedula=cedula, conjunto=conjunto).exists():
                    skipped_count += 1
                    return
                password = _DEFAULT_PASSWORD
                torre_nombre = str(row_data.get('agrupacion', '') or row_data.get('torre', '') or '').strip()
                torre_obj = torres_map.get(torre_nombre)
                apartamento = str(row_data.get('unidad', '') or row_data.get('apartamento', '') or '').strip()
                # Normalizar a 4 dígitos si es numérico (ej: "101" → "0101")
                if apartamento.isdigit() and len(apartamento) < 4:
                    apartamento = apartamento.zfill(4)
                Usuario.objects.create_user(
                    cedula=cedula,
                    nombre=nombre,
                    email=email,
                    password=password,
                    conjunto=conjunto,
                    user_type=user_type,
                    phone_number=str(row_data.get('telefono', '') or ''),
                    torre=torre_obj,
                    apartamento=apartamento,
                    must_change_password=True,
                )
                created_count += 1
                if send_emails and email:
                    result = _send_welcome(email, nombre, conjunto.nombre, cedula, password)
                    if result is not True:
                        email_errors.append(f'{email}: {result}')

            for sheet, utype in [('Administrador', 'administrador'), ('Propietarios', 'propietario'), ('Portería', 'porteria')]:
                ws_s = wb[sheet]
                hdrs = [c.value for c in next(ws_s.iter_rows(min_row=1, max_row=1))]
                for row in ws_s.iter_rows(min_row=2, values_only=True):
                    if not row[0]:
                        continue
                    create_user(dict(zip(hdrs, row)), utype)

            # Parqueaderos — hojas opcionales
            from kislevsmart.models import ParqueaderoCarro, ParqueaderoMoto
            for sheet_name, model in [('Parqueadero Carros', ParqueaderoCarro), ('Parqueadero Motos', ParqueaderoMoto)]:
                if sheet_name in wb.sheetnames:
                    ws_pk = wb[sheet_name]
                    pk_data = {row[0].value: row[1].value for row in ws_pk.iter_rows(min_row=2) if row[0].value}
                    espacios = int(pk_data.get('total_espacios', 0) or 0)
                    if espacios > 0:
                        model.objects.get_or_create(conjunto=conjunto, defaults={'total_espacios': espacios})

    except ValueError as exc:
        messages.error(request, str(exc))
        return render(request, 'accounts/upload_conjunto.html')
    except Exception as exc:
        messages.error(request, f'Error durante la importación: {exc}')
        return render(request, 'accounts/upload_conjunto.html')

    success_msg = f'Conjunto "{conjunto.nombre}" importado: {created_count} usuarios creados, {skipped_count} omitidos.'
    messages.success(request, success_msg)
    if email_errors:
        # Mostrar hasta 5 errores en detalle para diagnóstico
        detalle = ' | '.join(email_errors[:5])
        if len(email_errors) > 5:
            detalle += f' ... y {len(email_errors) - 5} más'
        messages.warning(request, f'{len(email_errors)} correos no enviados: {detalle}')
    return redirect('accounts:saas_dashboard')


# ── Gestión de usuarios (admin panel) ───────────────────────────────────────

from django.http import JsonResponse
from django.views.decorators.http import require_POST
import json

@login_required
def gestion_usuarios(request):
    """Panel de gestión de residentes para administradores (y saas_owner con ?conjunto=id)."""
    user = request.user

    if user.is_saas_owner and request.GET.get('conjunto'):
        try:
            conjunto = ConjuntoResidencial.objects.get(pk=request.GET['conjunto'])
        except ConjuntoResidencial.DoesNotExist:
            conjunto = user.conjunto
    elif user.user_type == 'administrador' or user.is_saas_owner:
        conjunto = user.conjunto
    else:
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden()

    from accounts.models import Torre
    torres = Torre.objects.filter(conjunto=conjunto, activo=True).order_by('nombre')

    torre_id = request.GET.get('torre')
    user_type_filter = request.GET.get('tipo', '')
    search = request.GET.get('q', '').strip()
    estado_filter = request.GET.get('estado', '')

    usuarios_qs = Usuario.objects.filter(conjunto=conjunto).exclude(is_saas_owner=True).order_by('torre__nombre', 'apartamento', 'nombre')

    if torre_id:
        usuarios_qs = usuarios_qs.filter(torre_id=torre_id)
    if user_type_filter == 'arrendatario':
        usuarios_qs = usuarios_qs.filter(user_type='propietario', es_arrendatario=True)
    elif user_type_filter:
        usuarios_qs = usuarios_qs.filter(user_type=user_type_filter, es_arrendatario=False)
    if search:
        from django.db.models import Q
        usuarios_qs = usuarios_qs.filter(Q(nombre__icontains=search) | Q(cedula__icontains=search) | Q(apartamento__icontains=search))
    if estado_filter == 'activo':
        usuarios_qs = usuarios_qs.filter(is_active=True)
    elif estado_filter == 'inactivo':
        usuarios_qs = usuarios_qs.filter(is_active=False)

    context = {
        'conjunto': conjunto,
        'torres': torres,
        'usuarios': usuarios_qs,
        'torre_id': torre_id,
        'user_type_filter': user_type_filter,
        'search': search,
        'estado_filter': estado_filter,
        'total': usuarios_qs.count(),
        'is_saas_owner': user.is_saas_owner,
        'all_conjuntos': ConjuntoResidencial.objects.all() if user.is_saas_owner else None,
        'conjunto_id': conjunto.pk,
    }
    return render(request, 'accounts/gestion_usuarios.html', context)


@require_POST
@login_required
def toggle_usuario_activo(request, usuario_id):
    """AJAX: activa o desactiva un usuario."""
    if request.user.user_type not in ('administrador',) and not request.user.is_saas_owner:
        return JsonResponse({'error': 'Sin permiso'}, status=403)
    try:
        target = Usuario.objects.get(pk=usuario_id)
    except Usuario.DoesNotExist:
        return JsonResponse({'error': 'No encontrado'}, status=404)
    if not request.user.is_saas_owner and target.conjunto != request.user.conjunto:
        return JsonResponse({'error': 'Sin permiso'}, status=403)
    target.is_active = not target.is_active
    target.save(update_fields=['is_active'])
    return JsonResponse({'is_active': target.is_active})


@require_POST
@login_required
def editar_usuario(request, usuario_id):
    """AJAX: edita torre, apartamento, teléfono y tipo de un usuario."""
    if request.user.user_type not in ('administrador',) and not request.user.is_saas_owner:
        return JsonResponse({'error': 'Sin permiso'}, status=403)
    try:
        target = Usuario.objects.get(pk=usuario_id)
    except Usuario.DoesNotExist:
        return JsonResponse({'error': 'No encontrado'}, status=404)
    if not request.user.is_saas_owner and target.conjunto != request.user.conjunto:
        return JsonResponse({'error': 'Sin permiso'}, status=403)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'JSON inválido'}, status=400)

    from accounts.models import Torre
    fields_updated = []

    if 'torre_id' in data:
        if data['torre_id']:
            try:
                torre = Torre.objects.get(pk=data['torre_id'], conjunto=target.conjunto)
                target.torre = torre
            except Torre.DoesNotExist:
                return JsonResponse({'error': 'Torre no válida'}, status=400)
        else:
            target.torre = None
        fields_updated.append('torre')

    if 'apartamento' in data:
        target.apartamento = str(data['apartamento']).strip()[:10]
        fields_updated.append('apartamento')

    if 'phone_number' in data:
        target.phone_number = str(data['phone_number']).strip()[:15]
        fields_updated.append('phone_number')

    if 'user_type' in data:
        tipo = data['user_type']
        if tipo == 'arrendatario':
            target.user_type = 'propietario'
            target.es_arrendatario = True
            fields_updated += ['user_type', 'es_arrendatario']
        elif tipo in ('propietario', 'administrador', 'porteria'):
            target.user_type = tipo
            target.es_arrendatario = False
            fields_updated += ['user_type', 'es_arrendatario']

    if 'email' in data:
        email_val = str(data['email']).strip()[:254]
        if email_val:
            target.email = email_val
            fields_updated.append('email')

    if fields_updated:
        target.save(update_fields=fields_updated)

    tipo_display = 'arrendatario' if target.es_arrendatario else target.user_type
    return JsonResponse({
        'ok': True,
        'torre': target.torre.nombre if target.torre else '',
        'apartamento': target.apartamento,
        'phone_number': target.phone_number or '',
        'user_type': tipo_display,
        'email': target.email or '',
    })


@require_POST
@login_required
def crear_usuario(request):
    """AJAX: crea un usuario en el conjunto del administrador (o el indicado para saas_owner)."""
    user = request.user
    if user.user_type not in ('administrador',) and not user.is_saas_owner:
        return JsonResponse({'error': 'Sin permiso'}, status=403)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'JSON inválido'}, status=400)

    # Determinar conjunto
    if user.is_saas_owner and data.get('conjunto_id'):
        try:
            conjunto = ConjuntoResidencial.objects.get(pk=data['conjunto_id'])
        except ConjuntoResidencial.DoesNotExist:
            return JsonResponse({'error': 'Conjunto no encontrado'}, status=404)
    else:
        conjunto = user.conjunto

    cedula   = str(data.get('cedula', '')).strip()
    nombre   = str(data.get('nombre', '')).strip()
    email    = str(data.get('email', '')).strip()
    telefono = str(data.get('telefono', '')).strip()[:15]
    tipo     = data.get('tipo', 'propietario')
    torre_id = data.get('torre_id')
    apartamento = str(data.get('apartamento', '')).strip()[:10]

    if not cedula or not nombre or not email:
        return JsonResponse({'error': 'Cédula, nombre y email son obligatorios'}, status=400)

    if Usuario.objects.filter(cedula=cedula, conjunto=conjunto).exists():
        return JsonResponse({'error': f'Ya existe un usuario con cédula {cedula} en este conjunto'}, status=400)

    # Normalizar apartamento a 4 dígitos si es numérico
    if apartamento and apartamento.isdigit() and len(apartamento) < 4:
        apartamento = apartamento.zfill(4)

    from accounts.models import Torre
    torre_obj = None
    if torre_id:
        try:
            torre_obj = Torre.objects.get(pk=torre_id, conjunto=conjunto)
        except Torre.DoesNotExist:
            pass

    es_arrendatario = tipo == 'arrendatario'
    user_type = 'propietario' if tipo in ('propietario', 'arrendatario') else tipo

    if user_type not in ('propietario', 'administrador', 'porteria'):
        return JsonResponse({'error': 'Tipo de usuario no válido'}, status=400)

    # ── Cédula en otro conjunto: sincronizar credenciales ────────────────────
    # Si la persona ya existe en otro conjunto usamos su hash actual para que
    # pueda iniciar sesión sin cambiar contraseña de nuevo.
    existing_other = Usuario.objects.filter(cedula=cedula).exclude(conjunto=conjunto).first()
    credenciales_sincronizadas = existing_other is not None

    nuevo = Usuario.objects.create_user(
        cedula=cedula,
        nombre=nombre,
        email=email,
        password=_DEFAULT_PASSWORD,
        conjunto=conjunto,
        user_type=user_type,
        es_arrendatario=es_arrendatario,
        phone_number=telefono or None,
        torre=torre_obj,
        apartamento=apartamento,
        must_change_password=not credenciales_sincronizadas,
    )

    if credenciales_sincronizadas:
        # Copiar hash — mismo hash, sin revelar la contraseña
        nuevo.password = existing_other.password
        nuevo.save(update_fields=['password'])

    # ── Enviar email de bienvenida con credenciales ───────────────────────────
    email_error = None
    try:
        from django.core.mail import EmailMultiAlternatives
        from django.template.loader import render_to_string
        login_url = getattr(settings, 'SITE_URL', 'https://kislev.net.co') + '/accounts/login/'
        if credenciales_sincronizadas:
            # Usuario ya tenía cuenta en otro conjunto — usar sus credenciales existentes
            text = (
                f"Hola {nombre},\n\nHas sido registrado en {conjunto.nombre} en Kislev.\n\n"
                f"Usuario: {cedula}\nContraseña: la misma que usas en tus otros conjuntos.\n\n"
                f"{login_url}"
            )
            msg = EmailMultiAlternatives(
                subject=f'Acceso a {conjunto.nombre} — Kislev',
                body=text,
                from_email=settings.DEFAULT_FROM_EMAIL,
                to=[email],
            )
        else:
            context = {
                'nombre': nombre,
                'conjunto_nombre': conjunto.nombre,
                'cedula': cedula,
                'password': _DEFAULT_PASSWORD,
                'login_url': login_url,
            }
            html = render_to_string('emails/bienvenida_credenciales.html', context)
            text = (
                f"Hola {nombre},\n\nHas sido registrado en {conjunto.nombre} en Kislev.\n\n"
                f"Usuario: {cedula}\nContraseña temporal: {_DEFAULT_PASSWORD}\n\n"
                f"Cámbiala en tu primer inicio de sesión.\n{login_url}"
            )
            msg = EmailMultiAlternatives(
                subject=f'Bienvenido a Kislev — {conjunto.nombre}',
                body=text,
                from_email=settings.DEFAULT_FROM_EMAIL,
                to=[email],
            )
            msg.attach_alternative(html, 'text/html')
        send_email_async(msg, detalle=f'Bienvenida → {email}')
        _log_envio_global('email', conjunto=conjunto, detalle=f'Bienvenida: {nombre}')
    except Exception as exc:
        email_error = str(exc)

    tipo_display = 'arrendatario' if es_arrendatario else user_type
    resp = {
        'ok': True,
        'id': nuevo.pk,
        'nombre': nuevo.nombre,
        'cedula': nuevo.cedula,
        'email': nuevo.email,
        'email_enviado': email_error is None,
        'email_error': email_error,
        'user_type': tipo_display,
        'torre': torre_obj.nombre if torre_obj else '',
        'apartamento': apartamento,
        'is_active': True,
        'credenciales_sincronizadas': credenciales_sincronizadas,
    }
    if credenciales_sincronizadas:
        resp['aviso'] = (
            f'Esta cédula ya existe en otro conjunto — las credenciales han sido sincronizadas. '
            f'El usuario puede iniciar sesión con su contraseña actual.'
        )
    return JsonResponse(resp)


@require_POST
@login_required
def toggle_conjunto_activo(request, conjunto_id):
    """AJAX: activa o desactiva un conjunto completo."""
    if not request.user.is_saas_owner:
        return JsonResponse({'error': 'Sin permiso'}, status=403)
    try:
        conjunto = ConjuntoResidencial.objects.get(pk=conjunto_id)
    except ConjuntoResidencial.DoesNotExist:
        return JsonResponse({'error': 'No encontrado'}, status=404)
    conjunto.estado = not conjunto.estado
    conjunto.save(update_fields=['estado'])
    return JsonResponse({'estado': conjunto.estado})


@require_POST
@login_required
def eliminar_conjunto(request, conjunto_id):
    """AJAX: elimina un conjunto y toda su data (CASCADE). Solo saas_owner."""
    if not request.user.is_saas_owner:
        return JsonResponse({'error': 'Sin permiso'}, status=403)
    try:
        conjunto = ConjuntoResidencial.objects.get(pk=conjunto_id)
    except ConjuntoResidencial.DoesNotExist:
        return JsonResponse({'error': 'Conjunto no encontrado'}, status=404)
    nombre = conjunto.nombre
    conjunto.delete()
    return JsonResponse({'ok': True, 'nombre': nombre})


@login_required
def exportar_usuarios_excel(request):
    """Descarga la lista de usuarios del conjunto en Excel."""
    if request.user.user_type not in ('administrador',) and not request.user.is_saas_owner:
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden()

    if request.user.is_saas_owner and request.GET.get('conjunto'):
        try:
            conjunto = ConjuntoResidencial.objects.get(pk=request.GET['conjunto'])
        except ConjuntoResidencial.DoesNotExist:
            conjunto = request.user.conjunto
    else:
        conjunto = request.user.conjunto

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        from django.http import HttpResponse
        return HttpResponse('openpyxl no instalado', status=500)

    import io
    from django.http import HttpResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Residentes'

    headers = ['Nombre', 'Cédula', 'Email', 'Teléfono', 'Torre', 'Apartamento', 'Tipo', 'Estado']
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', fgColor='4F2984')
    center = Alignment(horizontal='center')

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(h) + 4, 16)

    usuarios = Usuario.objects.filter(conjunto=conjunto).exclude(is_saas_owner=True).order_by('torre__nombre', 'apartamento', 'nombre')
    for row_idx, u in enumerate(usuarios, start=2):
        ws.cell(row=row_idx, column=1, value=u.nombre)
        ws.cell(row=row_idx, column=2, value=u.cedula)
        ws.cell(row=row_idx, column=3, value=u.email)
        ws.cell(row=row_idx, column=4, value=u.phone_number or '')
        ws.cell(row=row_idx, column=5, value=u.torre.nombre if u.torre else '')
        ws.cell(row=row_idx, column=6, value=u.apartamento)
        ws.cell(row=row_idx, column=7, value=u.get_user_type_display())
        ws.cell(row=row_idx, column=8, value='Activo' if u.is_active else 'Inactivo')

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="residentes_{conjunto.nit}.xlsx"'
    return response